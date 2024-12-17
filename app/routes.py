import os
import shutil
import sqlite3
import subprocess
import threading
from datetime import datetime, timezone
from flask import render_template, flash, url_for, request, redirect, send_file, session, jsonify, \
    render_template_string
from flask_login import current_user, login_user, logout_user, login_required
import sqlalchemy as sa
from urllib.parse import urlsplit, urlparse
from flask import Response
import openpyxl
from io import BytesIO

from pytz import utc
import json
from werkzeug.utils import secure_filename

from app import app, db
from app.forms import LoginForm, RegistrationForm, EditProfileForm, ProjectForm, TaskForm, UserForm, RoleForm, \
    ProjectStatisticsForm, StatusForm, PriorityForm, ChatMessageForm, ChatRoomForm, UserEditForm
from app.models import User, Project, Task, Role, Status, Priority, ProjectStatistics, ChatMessage, ChatRoom, \
    project_executors

# Путь к файлу логов
log_file_path = 'logs/microblog.log'


# Функция чтения логов
def read_logs(file_path):
    try:
        with open(file_path, 'r') as file:
            logs = file.readlines()
            return logs
    except FileNotFoundError:
        return ["Файл логов не найден"]


@app.route('/logs')
def show_logs():
    logs = read_logs(log_file_path)
    return render_template('logs.html', logs=logs)


# Главная страница приложения
@app.route('/')
def preindex():
    return render_template('preindex.html', title='Добро пожаловать!', is_login=True)


@app.route('/index')
@login_required
def index():
    projects = Project.query.filter_by(manager_id=current_user.id).all()
    return render_template('index.html', projects=projects)


@app.route('/admin_page')
@login_required
def admin_page():
    if not current_user.has_role('admin'):
        flash('У вас нет доступа к этой функции', 'error')
        return redirect(url_for('index'))
    return render_template('admin_page.html', title='Админ Панель')


# Профиль пользователя
@app.route('/user/<username>')
@login_required
def user(username):
    user = db.first_or_404(sa.select(User).where(User.username == username))
    projects = Project.query.filter_by(manager_id=user.id).all()
    return render_template('user.html', title='Профиль', user=user, projects=projects)


# Авторизация
@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:  # Проверяем, аутентифицирован ли пользователь
        if hasattr(current_user, 'has_role') and current_user.has_role('admin'):
            flash('Вы авторизовались как администратор!', 'success')
            return redirect(url_for('admin_page'))
        return redirect(url_for('index'))

    form = LoginForm()
    if form.validate_on_submit():
        user = db.session.scalar(
            sa.select(User).where(User.username == form.username.data)
        )
        if user is None or not user.check_password(form.password.data):
            flash('Неверное имя или пароль')
            return redirect(url_for('login'))

        login_user(user, remember=form.remember_me.data)

        # Логика обработки next_page
        next_page = request.args.get('next')
        if next_page:
            parsed_url = urlparse(next_page)
            # Проверяем, что next_page является относительным URL (внутри сайта)
            if not parsed_url.netloc and not parsed_url.scheme:
                return redirect(next_page)

        # Если next_page отсутствует или некорректен, перенаправляем на главную страницу
        if user.has_role('admin'):
            flash('Вы авторизовались как администратор!', 'success')
            return redirect(url_for('admin_page'))
        return redirect(url_for('index'))

    return render_template('login.html', title='Авторизация', form=form, is_login=True)


# Регистрация
@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('index'))

    form = RegistrationForm()
    if form.validate_on_submit():
        # Создаем нового пользователя
        user = User(username=form.username.data, email=form.email.data)
        user.set_password(form.password.data)

        # Назначаем роль "executor"
        executor_role = Role.query.filter_by(name="executor").first()
        if executor_role:
            user.roles.append(executor_role)
        else:
            flash('Роль "executor" не найдена. Обратитесь к администратору.', 'danger')
            return redirect(url_for('register'))

        # Сохраняем пользователя в базе данных
        db.session.add(user)
        db.session.commit()

        flash('Поздравляю, теперь вы зарегистрированный пользователь!')
        return redirect(url_for('login'))
    return render_template('register.html', title='Регистрация', form=form, is_login=True)


UPLOAD_FOLDER = os.path.join(app.static_folder, 'uploads')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER


@app.route('/edit_profile', methods=['GET', 'POST'])
@login_required
def edit_profile():
    form = EditProfileForm(original_username=current_user.username)

    if form.validate_on_submit():
        # Проверяем, если имя пользователя изменено
        if form.username.data and form.username.data != current_user.username:
            user = User.query.filter_by(username=form.username.data).first()
            if user:
                flash('Имя пользователя уже занято.', 'danger')
                return render_template('edit_profile.html', form=form, user=current_user)

        # Обновляем поля профиля
        current_user.username = form.username.data.strip() if form.username.data.strip() else None
        current_user.about_me = form.about_me.data.strip() if form.about_me.data.strip() else None
        current_user.first_name = form.first_name.data.strip() if form.first_name.data.strip() else None
        current_user.last_name = form.last_name.data.strip() if form.last_name.data.strip() else None
        current_user.phone_number = form.phone_number.data.strip() if form.phone_number.data.strip() else None
        current_user.date_birth = form.date_birth.data or None
        current_user.notify_birthday = form.notify_birthday.data

        # Обработка загрузки аватарки
        if form.avatar.data:
            avatar_file = form.avatar.data
            filename = secure_filename(f"{current_user.id}_{avatar_file.filename}")
            upload_folder = os.path.join('app', 'static', 'uploads', 'avatars')
            os.makedirs(upload_folder, exist_ok=True)
            filepath = os.path.join(upload_folder, filename)

            try:
                avatar_file.save(filepath)
                current_user.avatar_url = f"static/uploads/avatars/{filename}"
            except Exception as e:
                flash(f'Ошибка при сохранении аватара: {e}', 'danger')

        # Обработка переключения на Gravatar
        if form.use_gravatar.data:
            current_user.avatar_url = None  # Сбрасываем пользовательскую аватарку

        db.session.commit()
        flash('Изменения сохранены!', 'success')
        return redirect(url_for('user', username=current_user.username))

    elif request.method == 'GET':
        # Предзаполняем форму текущими данными
        form.username.data = current_user.username
        form.about_me.data = current_user.about_me
        form.first_name.data = current_user.first_name
        form.last_name.data = current_user.last_name
        form.phone_number.data = current_user.phone_number
        form.date_birth.data = current_user.date_birth
        form.use_gravatar.data = current_user.avatar_url is None

    return render_template('edit_profile.html', form=form, user=current_user)


# Выход из аккаунта
@app.route('/logout')
def logout():
    logout_user()
    flash('Вы успешно вышли из аккаунта.', 'success')
    return redirect(url_for('preindex'))


# Обновление времени последнего действия пользователя
@app.before_request
def before_request():
    if current_user.is_authenticated:
        current_user.last_seen = datetime.now(timezone.utc)
        db.session.commit()


@app.route('/projects')
@login_required
def projects():
    def get_avatar(user_id):
        user = db.session.query(User).filter_by(id=user_id).first()
        if user:
            return user.avatar_url or user.avatar(30)
        return "/static/default_avatar.png"

    # Проверяем роль пользователя
    if current_user.has_role('admin'):
        # Администратор видит все проекты
        projects_query = db.session.query(Project)
    else:
        # Остальные видят только свои проекты
        projects_query = (
            db.session.query(Project)
            .join(project_executors, project_executors.c.project_id == Project.id, isouter=True)
            .filter(
                (Project.manager_id == current_user.id) |
                (Project.responsible_id == current_user.id) |
                (project_executors.c.user_id == current_user.id)
            )
        )

    # Фильтрация по поисковому запросу
    search_query = request.args.get('search', '').strip()
    if search_query:
        projects_query = projects_query.filter(Project.title.ilike(f"%{search_query}%"))

    # Сортировка
    sort_option = request.args.get('sort', '')
    if sort_option == 'title_asc':
        projects_query = projects_query.order_by(Project.title.asc())
    elif sort_option == 'title_desc':
        projects_query = projects_query.order_by(Project.title.desc())
    elif sort_option == 'date_asc':
        projects_query = projects_query.order_by(Project.end_date.asc())
    elif sort_option == 'date_desc':
        projects_query = projects_query.order_by(Project.end_date.desc())
    elif sort_option == 'priority_asc':
        projects_query = projects_query.order_by(Project.priority_id.asc())
    elif sort_option == 'priority_desc':
        projects_query = projects_query.order_by(Project.priority_id.desc())

    # Преобразование результатов запроса
    projects_query = projects_query.all()
    projects = [
        {
            "id": project.id,
            "title": project.title,
            "icon_color": project.icon_color,
            "description": project.description,
            "end_date": project.end_date,
            "manager": {
                "first_name": project.manager.first_name,
                "last_name": project.manager.last_name,
                "username": project.manager.username,
                "avatar": get_avatar(project.manager_id),
            },
            "responsible_user": {
                "first_name": project.responsible_user.first_name if project.responsible_user else None,
                "last_name": project.responsible_user.last_name if project.responsible_user else None,
                "username": project.responsible_user.username if project.responsible_user else None,
                "avatar": get_avatar(project.responsible_id) if project.responsible_id else None,
            },
            "status_name": project.status.name if project.status else "Без статуса",
            "priority_level": project.priority.level if project.priority else "Не задан",
        }
        for project in projects_query
    ]

    # Получение задач по проектам
    project_tasks = {
        project["id"]: Task.query.filter(Task.project_id == project["id"]).all()
        for project in projects
    }

    return render_template('projects.html', projects=projects, project_tasks=project_tasks)


# Страница проекта
@app.route('/project/<int:project_id>')
@login_required
def project_detail(project_id):
    project = Project.query.get_or_404(project_id)

    # Проверка прав: администратор, управляющий, ответственный или участник проекта
    if not (
            current_user.has_role('admin') or
            project.manager_id == current_user.id or
            project.responsible_id == current_user.id or
            current_user in project.executors
    ):
        flash("У Вас нет доступа к данному проекту!")
        return redirect(url_for('index'))

    # Получаем задачи проекта
    tasks = Task.query.filter_by(project_id=project_id).all()

    # Форматирование даты начала
    start_date_utc = (
        project.start_date.astimezone(utc).strftime('%Y-%m-%d %H:%M:%S') if project.start_date else None
    )

    # Попытка получить статистику проекта
    project_stat = ProjectStatistics.query.filter_by(project_id=project_id).order_by(
        ProjectStatistics.updated_at.desc()).first()

    if project_stat:
        # Используем сохраненные данные статистики
        completed_tasks = project_stat.tasks_completed
        total_tasks = project_stat.tasks_total
    else:
        # Если статистики нет, рассчитываем данные
        completed_tasks = sum(1 for task in tasks if task.status and task.status.name == "Выполнено")
        total_tasks = len(tasks)

    # Рендер страницы проекта
    return render_template(
        'project_detail.html',
        project=project,
        tasks=tasks,
        completed_tasks=completed_tasks,
        total_tasks=total_tasks,
        start_date_utc=start_date_utc
    )


@app.route('/get_status_options')
@login_required
def get_status_options():
    statuses = Status.query.all()
    options = [{'id': status.id, 'name': status.name} for status in statuses]
    return jsonify({'options': options})


@app.route('/get_priority_options')
@login_required
def get_priority_options():
    priorities = Priority.query.all()
    options = [{'id': priority.id, 'name': priority.level} for priority in priorities]
    return jsonify({'options': options})


from datetime import datetime


@app.route('/project/<int:project_id>/update_field', methods=['POST'])
@login_required
def update_project_field(project_id):
    project = Project.query.get_or_404(project_id)

    # Получение данных из запроса
    data = request.get_json()
    if not data:
        return jsonify({'error': 'Данные не были отправлены.'}), 400

    print("Полученные данные:", data)  # Логируем входящие данные

    field_name = list(data.keys())[0]
    new_value = data[field_name]

    # Проверка прав для изменения полей
    if field_name in ['title', 'description']:
        if not current_user.has_role('admin') and project.manager_id != current_user.id:
            return jsonify({'error': 'У вас нет прав для изменения этого поля.'}), 403
    elif field_name in ['status', 'priority', 'end_date', 'responsible_user']:
        if not current_user.has_role(
                'admin') and project.manager_id != current_user.id and project.responsible_id != current_user.id:
            return jsonify({'error': 'У вас нет прав для изменения этого поля.'}), 403
    else:
        return jsonify({'error': 'Недопустимое поле для изменения.'}), 400

    # Логика изменения полей
    try:
        if field_name == 'title':
            if not new_value.strip():
                return jsonify({'error': 'Название проекта не может быть пустым.'}), 400
            project.title = new_value.strip()
            print(f"Обновлено название проекта: {project.title}")

        elif field_name == 'description':
            project.description = new_value.strip()
            print(f"Обновлено описание проекта: {project.description}")

        elif field_name == 'status':
            new_status = Status.query.get(new_value)
            if not new_status:
                return jsonify({'error': 'Указанный статус не найден.'}), 404
            old_status = project.status.name
            project.status_id = new_status.id

            # Установка даты начала, если статус меняется на "В процессе"
            if old_status == "Не начато" and new_status.name == "В процессе" and not project.start_date:
                project.start_date = datetime.now(timezone.utc)

            # Очистка дат, если статус возвращается на "Не начато"
            elif new_status.name == "Не начато":
                project.start_date = None
                project.end_date = None
            print(f"Обновлен статус проекта: {new_status.name}")

        elif field_name == 'priority':
            new_priority = Priority.query.get(new_value)
            if not new_priority:
                return jsonify({'error': 'Указанный приоритет не найден.'}), 404
            project.priority_id = new_priority.id
            print(f"Обновлен приоритет проекта: {new_priority.level}")

        elif field_name == 'end_date':
            if new_value:
                try:
                    end_date = datetime.strptime(new_value, '%Y-%m-%d')
                    project.end_date = end_date

                    # Если статус "Не начато", то меняем его на "В процессе" и устанавливаем дату начала
                    if project.status.name == "Не начато":
                        in_progress_status = Status.query.filter_by(name="В процессе").first()
                        if in_progress_status:
                            project.status_id = in_progress_status.id
                            project.start_date = datetime.now(timezone.utc)
                    print(f"Обновлена дата окончания: {project.end_date}")
                except ValueError:
                    return jsonify({'error': 'Неверный формат даты. Используйте YYYY-MM-DD.'}), 400
            else:
                project.end_date = None
                print("Дата окончания проекта очищена")

        elif field_name == 'responsible_user':
            new_responsible_user = User.query.get(new_value)
            if not new_responsible_user:
                return jsonify({'error': 'Указанный пользователь не найден.'}), 404
            project.responsible_id = new_responsible_user.id
            print(f"Обновлен ответственный: {new_responsible_user.first_name} {new_responsible_user.last_name}")

        # Сохранение изменений в базе данных
        db.session.commit()
        print("Изменения успешно сохранены в базе данных.")

    except Exception as e:
        print("Ошибка при обработке изменений:", str(e))
        return jsonify({'error': 'Произошла ошибка при обновлении.'}), 500

    # Подготовка данных для ответа
    response_data = {
        'success': True,
        'title': project.title,
        'description': project.description,
        'start_date': project.start_date.strftime('%d.%m.%Y') if project.start_date else None,
        'end_date': project.end_date.strftime('%d.%m.%Y') if project.end_date else None,
        'status': project.status.name if project.status else "Без статуса",
        'priority': project.priority.level if project.priority else "Не задан",
        'responsible_user': f"{project.responsible_user.first_name} {project.responsible_user.last_name}" if project.responsible_user else "Не назначен"
    }

    return jsonify(response_data), 200


@app.route('/project/<int:project_id>/add_member', methods=['POST'])
@login_required
def add_project_member(project_id):
    project = Project.query.get_or_404(project_id)

    # Проверка прав
    if not (current_user.has_role(
            'admin') or project.manager_id == current_user.id or project.responsible_id == current_user.id):
        flash("У вас нет прав для добавления участников.")
        return redirect(url_for('project_detail', project_id=project_id))

    # Получение данных из формы
    user_email = request.form.get('user_email')
    user = User.query.filter_by(email=user_email).first()

    if not user:
        flash("Пользователь с указанным email не найден.")
        return redirect(url_for('project_detail', project_id=project_id))

    if user in project.executors:
        flash("Пользователь уже является участником проекта.")
        return redirect(url_for('project_detail', project_id=project_id))

    # Добавление участника
    project.executors.append(user)
    db.session.commit()

    flash(f"Пользователь {user_email} добавлен в проект.")
    return redirect(url_for('project_detail', project_id=project_id))


@app.route('/project/<int:project_id>/remove_member/<int:user_id>', methods=['POST'])
@login_required
def remove_project_member(project_id, user_id):
    project = Project.query.get_or_404(project_id)

    # Проверка прав
    if project.manager_id != current_user.id and project.responsible_id != current_user.id:
        return jsonify({'error': 'Нет прав для удаления участника.'}), 403

    user = User.query.get_or_404(user_id)

    if user not in project.executors:
        return jsonify({'error': 'Пользователь не является участником.'}), 404

    # Удаление участника
    project.executors.remove(user)
    db.session.commit()

    return jsonify({'success': True, 'message': f"Пользователь {user.username} удален из проекта."})


@app.route('/project/<int:project_id>/update_description', methods=['POST'])
@login_required
def update_project_description(project_id):
    project = Project.query.get_or_404(project_id)

    # Проверяем права на редактирование
    if not current_user.has_role('admin') and project.manager_id != current_user.id:
        return jsonify({"error": "У Вас нет разрешения редактировать данный проект!"}), 403

    # Получаем данные из запроса
    data = request.get_json()
    new_description = data.get('description', '').strip()

    # Если описание пустое, устанавливаем None
    project.description = new_description if new_description else None

    # Сохраняем изменения
    db.session.commit()

    return jsonify({"message": "Описание проекта обновлено", "success": True}), 200


@app.route('/project/<int:project_id>/update_title', methods=['POST'])
@login_required
def update_project_title(project_id):
    project = Project.query.get_or_404(project_id)

    # Проверяем права
    if not current_user.has_role('admin') and project.manager_id != current_user.id:
        return jsonify({'error': 'Нет доступа'}), 403

    data = request.get_json()
    new_title = data.get('title', '').strip()

    # Если пользователь удаляет название, устанавливаем "Новый проект"
    project.title = new_title if new_title else "Новый проект"

    # Сохраняем изменения
    db.session.commit()

    return jsonify({'message': 'Название обновлено', 'success': True, 'new_title': project.title}), 200


@app.route('/get_users', methods=['GET'])
@login_required
def get_users():
    users = User.query.all()
    user_options = [
        {'id': user.id, 'name': f"{user.first_name or ''} {user.last_name or ''}".strip() or user.username}
        for user in users
    ]
    return jsonify({'options': user_options})


# Страница задачи
@app.route('/task/<int:task_id>')
@login_required
def task_detail(task_id):
    task = Task.query.get_or_404(task_id)
    project = Project.query.get(task.project_id)
    if not (
            current_user.has_role('admin') or
            project.manager_id == current_user.id or
            project.responsible_id == current_user.id or
            current_user in project.executors
    ):
        flash("У Вас нет разрешения просматривать данную задачу")
        return redirect(url_for('index'))
    return render_template('task_detail.html', task=task)


@app.route('/admin/create_project', methods=['GET', 'POST'])
@login_required
def create_project_admin():
    form = ProjectForm()
    form.populate_choices()  # Загружаем данные в выборы формы
    if form.validate_on_submit():
        project = Project(
            title=form.title.data,
            description=form.description.data,
            status_id=form.status_id.data,
            priority_id=form.priority_id.data,
            manager_id=current_user.id,
            responsible_id=form.responsible_id.data
        )
        db.session.add(project)
        db.session.commit()
        flash("Проект успешно создан.")
        return redirect(url_for('projects'))
    return render_template('create_project.html', form=form)


import random


@app.route('/create_project', methods=['POST'])
@login_required
def create_project():
    # Получение статуса "Не начато" и приоритета "Низкий" из базы данных
    not_started_status = db.session.query(Status).filter_by(name="Не начато").first()
    low_priority = db.session.query(Priority).filter_by(level="Низкий").first()

    # Генерация случайного цвета для иконки проекта
    def generate_random_color():
        return f"#{random.randint(0, 255):02x}{random.randint(0, 255):02x}{random.randint(0, 255):02x}"

    # Создаем новый проект с заданными параметрами
    project = Project(
        title="Новый проект",
        description=None,
        manager_id=current_user.id,
        responsible_id=current_user.id,
        status_id=not_started_status.id if not_started_status else None,
        priority_id=low_priority.id if low_priority else None,
        icon_color=generate_random_color()  # Устанавливаем случайный цвет иконки
    )

    # Сохраняем проект в базе данных
    db.session.add(project)
    db.session.commit()

    # Уведомление и редирект на страницу деталей проекта
    flash("Проект успешно создан!", "success")
    return redirect(url_for('project_detail', project_id=project.id))


@app.route('/create_task/<int:project_id>', methods=['GET', 'POST'])
@login_required
def create_task(project_id):
    project = Project.query.get_or_404(project_id)
    if not current_user.has_role(
            'admin') and project.manager_id != current_user.id and project.responsible_id != current_user.id:
        flash("У вас нет разрешения добавлять задачи в данный проект!")
        return redirect(url_for('project_detail', project_id=project_id))
    form = TaskForm()
    form.populate_choices()
    if form.validate_on_submit():
        task = Task(
            title=form.title.data,
            description=form.description.data,
            project_id=project_id,
            status_id=form.status_id.data,
            priority_id=form.priority_id.data,
            deadline=form.deadline.data
        )
        db.session.add(task)
        db.session.commit()
        flash("Задача успешно создана.")
        return redirect(url_for('project_detail', project_id=project_id))
    return render_template('create_task.html', form=form, project=project)


def execute_command(command):
    try:
        result = subprocess.run(
            command, shell=True, check=True, capture_output=True, text=True, timeout=60
        )
        print(f"Command output: {result.stdout}")
        return result.stdout
    except subprocess.CalledProcessError as e:
        print(f"Command failed: {e.stderr}")
        raise
    except Exception as e:
        print(f"Unexpected error: {str(e)}")
        raise


@app.route('/admin/backup', methods=['GET'])
def backup_database():
    if not current_user.has_role('admin'):
        flash('У вас нет доступа к этой функции', 'error')
        return redirect(url_for('index'))

    backup_dir = os.path.join(os.getcwd(), "backups")
    os.makedirs(backup_dir, exist_ok=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_filename = f"backup_{timestamp}.sql"
    backup_path = os.path.join(backup_dir, backup_filename)

    try:
        db_url = app.config['SQLALCHEMY_DATABASE_URI']
        with open(backup_path, 'w') as f:
            f.write("DROP SCHEMA public CASCADE;\nCREATE SCHEMA public;\n")

        # Создание бэкапа
        subprocess.run(
            ['pg_dump', f'--dbname={db_url}'],
            stdout=open(backup_path, 'a'),
            stderr=subprocess.PIPE,
            check=True,
            text=True
        )
        session['backup_file'] = backup_filename
        flash('Бэкап успешно создан. Вы можете его скачать.', 'success')
    except subprocess.CalledProcessError as e:
        flash(f'Ошибка при создании бэкапа: {e.stderr}', 'error')
    except Exception as e:
        flash(f'Ошибка: {str(e)}', 'error')

    return redirect(url_for('admin_page'))


def reset_schema_and_restore(backup_path, db_url):
    try:
        print("Starting schema reset...")
        # Очистка схемы базы данных
        clear_schema_cmd = f'psql --dbname={db_url} -c "DROP SCHEMA public CASCADE; CREATE SCHEMA public;"'
        execute_command(clear_schema_cmd)
        print("Schema reset complete.")

        print("Restoring from backup...")
        # Восстановление базы данных из дампа
        restore_cmd = f'psql --dbname={db_url} --file={backup_path}'
        execute_command(restore_cmd)
        print("Restore completed successfully.")
    except Exception as e:
        print(f"Error during restore: {e}")


@app.route('/admin/upload_backup', methods=['POST'])
def upload_backup():
    if not current_user.has_role('admin'):
        flash('У вас нет доступа к этой функции', 'error')
        return redirect(url_for('index'))

    if 'file' not in request.files:
        flash('Нет файла в запросе', 'error')
        return redirect(url_for('index'))

    file = request.files['file']
    if file.filename == '':
        flash('Файл не выбран', 'error')
        return redirect(url_for('admin_page'))

    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        backup_path = os.path.join(os.getcwd(), "backups", filename)
        file.save(backup_path)
        print(f"Backup file saved at {backup_path}")

        db_url = app.config['SQLALCHEMY_DATABASE_URI']

        # Запускаем процесс восстановления в фоновом потоке
        thread = threading.Thread(target=reset_schema_and_restore, args=(backup_path, db_url))
        thread.start()

        flash('Восстановление запущено. Пожалуйста, подождите.', 'success')
        return redirect(url_for('admin_page'))

    flash('Неверный формат файла. Пожалуйста, загрузите файл .sql.', 'error')
    return redirect(url_for('admin_page'))


def allowed_file(filename):
    ALLOWED_EXTENSIONS = {'sql'}
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@app.route('/admin/download_backup', methods=['GET'])
def download_backup():
    backup_filename = session.pop('backup_file', None)
    if not backup_filename:
        flash('Файл бэкапа не найден', 'error')
        return redirect(url_for('admin_page'))

    backup_path = os.path.join(os.getcwd(), "backups", backup_filename)
    return send_file(backup_path, as_attachment=True)


def allowed_file(filename):
    ALLOWED_EXTENSIONS = {'sql'}
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# Отображение всех проектов для админ-панели
@app.route('/admin/projects')
@login_required
def admin_projects():
    projects = Project.query.all()
    return render_template('admin_projects.html', projects=projects)


# Редактирование проекта
@app.route('/admin/edit_project/<int:project_id>', methods=['GET', 'POST'])
@login_required
def edit_project(project_id):
    project = Project.query.get_or_404(project_id)
    form = ProjectForm(obj=project)
    form.populate_choices()
    if form.validate_on_submit():
        project.title = form.title.data
        project.description = form.description.data
        project.status_id = form.status_id.data
        project.priority_id = form.priority_id.data
        project.responsible_id = form.responsible_id.data
        db.session.commit()
        flash('Проект успешно обновлен!')
        return redirect(url_for('admin_projects'))
    return render_template('edit_project.html', form=form, project=project)


# ТРАНЗАКЦИЯ
from sqlalchemy.orm import scoped_session


@app.route('/delete_project', methods=['POST'])
@login_required
def delete_project():
    session = scoped_session(db.session)

    project_ids = request.form.getlist('project_ids')
    try:
        if project_ids:
            for project_id in project_ids:
                project = session.query(Project).get(project_id)
                if project:
                    for task in project.tasks:
                        session.delete(task)
                    session.delete(project)
        else:
            project_id = request.form.get('project_id')
            if project_id:
                project = session.query(Project).get_or_404(project_id)
                for task in project.tasks:
                    session.delete(task)
                session.delete(project)
            else:
                flash("Не выбрано ни одного проекта для удаления.", "error")
                return redirect(url_for('admin_projects'))

        session.commit()
        flash("Выбранные проекты и связанные задачи успешно удалены.", "success")
    except Exception as e:
        session.rollback()
        flash(f"Ошибка при удалении проекта: {str(e)}", "error")
    finally:
        session.remove()  # Закрытие локальной сессии

    return redirect(url_for('admin_projects'))


# Хранимая процедура

from sqlalchemy import text

@app.route('/create_project_with_tasks', methods=['POST'])
def create_project_with_tasks():
    try:
        data = request.get_json()
        project_title = data.get('title')
        project_description = data.get('description')
        tasks = data.get('tasks', [])

        if not project_title or not tasks:
            return jsonify({'error': 'Название проекта и задачи обязательны!'}), 400

        # Получаем ID авторизованного пользователя
        manager_id = 1

        # Преобразуем задачи в JSON
        tasks_json = json.dumps(tasks)

        # Определяем SQL-запрос как text
        query = text("""
        CALL create_project_with_tasks(:project_title, :project_description, :manager_id, :tasks)
        """)

        # Выполняем запрос
        db.session.execute(query, {
            'project_title': project_title,
            'project_description': project_description,
            'manager_id': manager_id,
            'tasks': tasks_json
        })
        db.session.commit()

        return jsonify({'message': 'Проект и задачи успешно созданы!'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Ошибка при выполнении: {str(e)}'}), 500



# Роут для отображения всех задач в админке
@app.route('/admin/tasks')
@login_required
def admin_tasks():
    tasks = Task.query.all()
    return render_template('admin_tasks.html', tasks=tasks)


# Роут для создания новой задачи в админке
@app.route('/admin/create_task', methods=['GET', 'POST'])
@login_required
def admin_create_task():
    form = TaskForm()
    form.populate_choices()
    if form.validate_on_submit():
        task = Task(
            title=form.title.data,
            description=form.description.data,
            status_id=form.status_id.data,
            priority_id=form.priority_id.data,
            deadline=form.deadline.data
        )
        db.session.add(task)
        db.session.commit()
        flash("Задача успешно создана.")
        return redirect(url_for('admin_tasks'))
    return render_template('create_task.html', form=form)


# Роут для редактирования задачи в админке
@app.route('/admin/edit_task/<int:task_id>', methods=['GET', 'POST'])
@login_required
def edit_task(task_id):
    task = Task.query.get_or_404(task_id)
    form = TaskForm(obj=task)
    form.populate_choices()
    if form.validate_on_submit():
        task.title = form.title.data
        task.description = form.description.data
        task.status_id = form.status_id.data
        task.priority_id = form.priority_id.data
        task.deadline = form.deadline.data
        db.session.commit()
        flash("Задача успешно обновлена.")
        return redirect(url_for('admin_tasks'))
    return render_template('edit_task.html', form=form, task=task)


@app.route('/admin/delete_task', methods=['POST'])
@login_required
def delete_task():
    task_ids = request.form.getlist('task_ids')  # Для множественного удаления
    task_id = request.form.get('task_id')  # Для одиночного удаления

    if task_ids:  # Множественное удаление
        for task_id in task_ids:
            task = Task.query.get(task_id)
            if task:
                db.session.delete(task)
        db.session.commit()
        flash("Выбранные задачи успешно удалены.", "success")
    elif task_id:  # Одиночное удаление
        task = Task.query.get_or_404(task_id)
        db.session.delete(task)
        db.session.commit()
        flash("Задача успешно удалена.", "success")
    else:  # Если данные не переданы
        flash("Не выбрано ни одной задачи для удаления.", "error")

    return redirect(url_for('admin_tasks'))


@app.route('/admin/create_user', methods=['GET', 'POST'])
@login_required
def create_user():
    form = UserForm()
    form.roles.choices = [(role.id, role.name) for role in Role.query.all()]  # Загрузка ролей

    if form.validate_on_submit():
        user = User(
            username=form.username.data,
            email=form.email.data,
            first_name=form.first_name.data,
            last_name=form.last_name.data,
            middle_name=form.middle_name.data,
            phone_number=form.phone_number.data,
            date_birth=form.date_birth.data,
            about_me=form.about_me.data,
            avatar=form.avatar.data,
            is_active=form.is_active.data
        )
        user.set_password(form.password.data)

        # Привязка выбранных ролей
        selected_roles = Role.query.filter(Role.id.in_(form.roles.data)).all()
        user.roles.extend(selected_roles)

        db.session.add(user)
        db.session.commit()
        flash("Пользователь успешно создан.")
        return redirect(url_for('admin_users'))
    return render_template('create_user.html', form=form)


# Отображение всех пользователей для админ-панели
@app.route('/admin/users')
@login_required
def admin_users():
    users = User.query.all()
    return render_template('admin_users.html', users=users)


@app.route('/admin/edit_user/<int:user_id>', methods=['GET', 'POST'])
@login_required
def edit_user(user_id):
    user = User.query.get_or_404(user_id)
    form = UserEditForm(obj=user)

    # Загрузка всех доступных ролей для выбора
    form.roles.choices = [(role.id, role.name) for role in Role.query.all()]

    # Установка текущих ролей пользователя в форме
    if request.method == 'GET':
        form.roles.data = [role.id for role in user.roles]
        form.is_active.data = user.is_active

    if form.validate_on_submit():
        user.username = form.username.data
        user.email = form.email.data
        user.first_name = form.first_name.data
        user.last_name = form.last_name.data
        user.middle_name = form.middle_name.data
        user.phone_number = form.phone_number.data
        user.date_birth = form.date_birth.data
        user.about_me = form.about_me.data
        user.avatar = form.avatar.data
        user.is_active = form.is_active.data

        # Обновление пароля, если он изменён
        if form.password.data:
            user.set_password(form.password.data)

        # Обновление ролей
        selected_roles = Role.query.filter(Role.id.in_(form.roles.data)).all()
        user.roles = selected_roles

        db.session.commit()
        flash('Пользователь успешно обновлен!')
        return redirect(url_for('admin_users'))
    return render_template('edit_user.html', form=form, user=user)


@app.route('/admin/delete_users', methods=['POST'])
@login_required
def delete_users():
    user_ids = request.form.getlist('user_ids')  # Для множественного удаления
    user_id = request.form.get('user_id')  # Для одиночного удаления

    if user_ids:  # Множественное удаление
        for user_id in user_ids:
            user = User.query.get(user_id)
            if user:
                db.session.delete(user)
        db.session.commit()
        flash("Выбранные пользователи успешно удалены.", "success")
    elif user_id:  # Одиночное удаление
        user = User.query.get_or_404(user_id)
        db.session.delete(user)
        db.session.commit()
        flash("Пользователь успешно удалён.", "success")
    else:  # Если данные не переданы
        flash("Не выбрано ни одного пользователя для удаления.", "error")

    return redirect(url_for('admin_users'))


# Создание новой роли
@app.route('/admin/create_role', methods=['GET', 'POST'])
@login_required
def create_role():
    form = RoleForm()
    if form.validate_on_submit():
        role = Role(name=form.name.data)
        db.session.add(role)
        db.session.commit()
        flash("Роль успешно создана.")
        return redirect(url_for('admin_roles'))
    return render_template('create_role.html', form=form)


# Отображение всех ролей для админ-панели
@app.route('/admin/roles')
@login_required
def admin_roles():
    roles = Role.query.all()
    return render_template('admin_roles.html', roles=roles)


# Редактирование роли
@app.route('/admin/edit_role/<int:role_id>', methods=['GET', 'POST'])
@login_required
def edit_role(role_id):
    role = Role.query.get_or_404(role_id)
    form = RoleForm(obj=role)
    if form.validate_on_submit():
        role.name = form.name.data
        db.session.commit()
        flash('Роль успешно обновлена!')
        return redirect(url_for('admin_roles'))
    return render_template('edit_role.html', form=form, role=role)


@app.route('/admin/delete_roles', methods=['POST'])
@login_required
def delete_role():
    role_ids = request.form.getlist('role_ids')  # Для множественного удаления
    role_id = request.form.get('role_id')  # Для одиночного удаления

    if role_ids:  # Если выбран хотя бы один чекбокс
        for role_id in role_ids:
            role = Role.query.get(role_id)
            if role:
                db.session.delete(role)
        db.session.commit()
        flash("Выбранные роли успешно удалены.", "success")
    elif role_id:  # Если удаляется одна роль
        role = Role.query.get_or_404(role_id)
        db.session.delete(role)
        db.session.commit()
        flash("Роль успешно удалена.", "success")
    else:
        flash("Не выбрано ни одной роли для удаления.", "error")

    return redirect(url_for('admin_roles'))


from datetime import datetime, timezone


@app.route('/admin/create_project_stat', methods=['GET', 'POST'])
@login_required
def create_project_stat():
    form = ProjectStatisticsForm()
    form.populate_choices()  # Заполняем выбор проектов
    if form.validate_on_submit():
        project_stat = ProjectStatistics(
            project_id=form.project_id.data,
            tasks_total=form.tasks_total.data,
            tasks_completed=form.tasks_completed.data,
            average_completion_time=form.average_completion_time.data,
            updated_at=datetime.now(timezone.utc)  # Устанавливаем значение updated_at вручную
        )
        db.session.add(project_stat)
        db.session.commit()
        flash("Статистика проекта успешно создана.")
        return redirect(url_for('admin_project_stats'))
    return render_template('create_project_stat.html', form=form)


# Отображение всех записей статистики для админ-панели
@app.route('/admin/project_stats')
@login_required
def admin_project_stats():
    project_stats = ProjectStatistics.query.all()
    return render_template('admin_project_stats.html', project_stats=project_stats)


# Редактирование записи статистики проекта
@app.route('/admin/edit_project_stat/<int:stat_id>', methods=['GET', 'POST'])
@login_required
def edit_project_stat(stat_id):
    project_stat = ProjectStatistics.query.get_or_404(stat_id)
    form = ProjectStatisticsForm(obj=project_stat)
    form.populate_choices()  # Заполняем выбор проектов
    if form.validate_on_submit():
        project_stat.project_id = form.project_id.data
        project_stat.tasks_total = form.tasks_total.data
        project_stat.tasks_completed = form.tasks_completed.data
        project_stat.average_completion_time = form.average_completion_time.data
        db.session.commit()
        flash('Статистика проекта успешно обновлена!')
        return redirect(url_for('admin_project_stats'))
    return render_template('edit_project_stat.html', form=form, project_stat=project_stat)


@app.route('/admin/delete_project_stats', methods=['POST'])
@login_required
def delete_project_stat():
    stat_ids = request.form.getlist('stat_ids')  # Для множественного удаления
    stat_id = request.form.get('stat_id')  # Для одиночного удаления

    if stat_ids:  # Множественное удаление
        for stat_id in stat_ids:
            project_stat = ProjectStatistics.query.get(stat_id)
            if project_stat:
                db.session.delete(project_stat)
        db.session.commit()
        flash("Выбранные записи статистики успешно удалены.", "success")
    elif stat_id:  # Одиночное удаление
        project_stat = ProjectStatistics.query.get_or_404(stat_id)
        db.session.delete(project_stat)
        db.session.commit()
        flash("Запись статистики успешно удалена.", "success")
    else:  # Если данные не переданы
        flash("Не выбрано ни одной записи для удаления.", "error")

    return redirect(url_for('admin_project_stats'))


@app.route('/admin/statuses')
@login_required
def admin_statuses():
    statuses = Status.query.all()
    return render_template('admin_statuses.html', statuses=statuses)


@app.route('/admin/create_status', methods=['GET', 'POST'])
@login_required
def create_status():
    form = StatusForm()
    if form.validate_on_submit():
        status = Status(name=form.name.data)
        db.session.add(status)
        db.session.commit()
        flash("Статус успешно создан.")
        return redirect(url_for('admin_statuses'))
    return render_template('create_status.html', form=form)


@app.route('/admin/edit_status/<int:status_id>', methods=['GET', 'POST'])
@login_required
def edit_status(status_id):
    status = Status.query.get_or_404(status_id)
    form = StatusForm(obj=status)
    if form.validate_on_submit():
        status.name = form.name.data
        db.session.commit()
        flash("Статус успешно обновлён.")
        return redirect(url_for('admin_statuses'))
    return render_template('edit_status.html', form=form, status=status)


@app.route('/delete_status', methods=['POST'])
@login_required
def delete_status():
    status_ids = request.form.getlist('status_ids')
    if status_ids:
        for status_id in status_ids:
            status = Status.query.get(status_id)
            if status:
                db.session.delete(status)
        db.session.commit()
        flash("Выбранные статусы успешно удалены.", "success")
    else:
        status_id = request.form.get('status_id')
        if status_id:
            status = Status.query.get_or_404(status_id)
            db.session.delete(status)
            db.session.commit()
            flash("Статус успешно удалён.", "success")
        else:
            flash("Не выбрано ни одного статуса для удаления.", "error")
    return redirect(url_for('admin_statuses'))


@app.route('/admin/priorities')
@login_required
def admin_priorities():
    priorities = Priority.query.all()
    return render_template('admin_priorities.html', priorities=priorities)


@app.route('/admin/create_priority', methods=['GET', 'POST'])
@login_required
def create_priority():
    form = PriorityForm()
    if form.validate_on_submit():
        priority = Priority(level=form.level.data)
        db.session.add(priority)
        db.session.commit()
        flash("Приоритет успешно создан.")
        return redirect(url_for('admin_priorities'))
    return render_template('create_priority.html', form=form)


@app.route('/admin/edit_priority/<int:priority_id>', methods=['GET', 'POST'])
@login_required
def edit_priority(priority_id):
    priority = Priority.query.get_or_404(priority_id)
    form = PriorityForm(obj=priority)
    if form.validate_on_submit():
        priority.level = form.level.data
        db.session.commit()
        flash("Приоритет успешно обновлён.")
        return redirect(url_for('admin_priorities'))
    return render_template('edit_priority.html', form=form, priority=priority)


@app.route('/delete_priority', methods=['POST'])
@login_required
def delete_priority():
    priority_ids = request.form.getlist('priority_ids')
    if priority_ids:
        for priority_id in priority_ids:
            priority = Priority.query.get(priority_id)
            if priority:
                db.session.delete(priority)
        db.session.commit()
        flash("Выбранные приоритеты успешно удалены.", "success")
    else:
        priority_id = request.form.get('priority_id')
        if priority_id:
            priority = Priority.query.get_or_404(priority_id)
            db.session.delete(priority)
            db.session.commit()
            flash("Приоритет успешно удалён.", "success")
        else:
            flash("Не выбрано ни одного приоритета для удаления.", "error")
    return redirect(url_for('admin_priorities'))


@app.route('/admin/chat_messages')
@login_required
def admin_chat_messages():
    chat_messages = ChatMessage.query.all()
    return render_template('admin_chat_messages.html', chat_messages=chat_messages)


@app.route('/admin/create_chat_message', methods=['GET', 'POST'])
@login_required
def create_chat_message():
    form = ChatMessageForm()
    form.populate_choices()
    if form.validate_on_submit():
        chat_message = ChatMessage(
            chat_room_id=form.chat_room_id.data,
            user_id=form.user_id.data,
            content=form.content.data
        )
        db.session.add(chat_message)
        db.session.commit()
        flash("Сообщение успешно создано.")
        return redirect(url_for('admin_chat_messages'))
    return render_template('create_chat_message.html', form=form)


@app.route('/admin/edit_chat_message/<int:message_id>', methods=['GET', 'POST'])
@login_required
def edit_chat_message(message_id):
    chat_message = ChatMessage.query.get_or_404(message_id)
    form = ChatMessageForm(obj=chat_message)
    form.populate_choices()
    if form.validate_on_submit():
        chat_message.chat_room_id = form.chat_room_id.data
        chat_message.user_id = form.user_id.data
        chat_message.content = form.content.data
        db.session.commit()
        flash("Сообщение успешно обновлено.")
        return redirect(url_for('admin_chat_messages'))
    return render_template('edit_chat_message.html', form=form, chat_message=chat_message)


@app.route('/delete_chat_message', methods=['POST'])
@login_required
def delete_chat_message():
    message_ids = request.form.getlist('message_ids')
    if message_ids:
        for message_id in message_ids:
            message = ChatMessage.query.get(message_id)
            if message:
                db.session.delete(message)
        db.session.commit()
        flash("Выбранные сообщения успешно удалены.", "success")
    else:
        message_id = request.form.get('message_id')
        if message_id:
            message = ChatMessage.query.get_or_404(message_id)
            db.session.delete(message)
            db.session.commit()
            flash("Сообщение успешно удалено.", "success")
        else:
            flash("Не выбрано ни одного сообщения для удаления.", "error")
    return redirect(url_for('admin_chat_messages'))


@app.route('/admin/chat_rooms')
@login_required
def admin_chat_rooms():
    chat_rooms = ChatRoom.query.all()
    return render_template('admin_chat_rooms.html', chat_rooms=chat_rooms)


@app.route('/admin/create_chat_room', methods=['GET', 'POST'])
@login_required
def create_chat_room():
    form = ChatRoomForm()
    form.populate_choices()

    # Проверка, что форма прошла валидацию
    if form.validate_on_submit():
        if form.choice.data == 'project' and not form.project_id.data:
            flash('Необходимо выбрать проект!', 'error')
            return redirect(url_for('create_chat_room'))

        if form.choice.data == 'task' and not form.task_id.data:
            flash('Необходимо выбрать задачу!', 'error')
            return redirect(url_for('create_chat_room'))

        # Создание чат-комнаты
        chat_room = ChatRoom(
            type=form.type.data
        )

        if form.choice.data == 'project':
            chat_room.project_id = form.project_id.data
        elif form.choice.data == 'task':
            chat_room.task_id = form.task_id.data

        db.session.add(chat_room)
        db.session.commit()
        flash("Чат-комната успешно создана.")
        return redirect(url_for('admin_chat_rooms'))

    # Если форма не прошла валидацию, вернемся к тому же шаблону
    return render_template('create_chat_room.html', form=form)


@app.route('/admin/edit_chat_room/<int:room_id>', methods=['GET', 'POST'])
@login_required
def edit_chat_room(room_id):
    chat_room = ChatRoom.query.get_or_404(room_id)
    form = ChatRoomForm(obj=chat_room)
    form.populate_choices()
    if form.validate_on_submit():
        chat_room.project_id = form.project_id.data
        chat_room.task_id = form.task_id.data
        chat_room.type = form.type.data
        db.session.commit()
        flash("Чат-комната успешно обновлена.")
        return redirect(url_for('admin_chat_rooms'))
    return render_template('edit_chat_room.html', form=form, chat_room=chat_room)


@app.route('/delete_chat_room', methods=['POST'])
@login_required
def delete_chat_room():
    room_ids = request.form.getlist('room_ids')
    if room_ids:
        for room_id in room_ids:
            room = ChatRoom.query.get(room_id)
            if room:
                db.session.delete(room)
        db.session.commit()
        flash("Выбранные комнаты успешно удалены.", "success")
    else:
        room_id = request.form.get('room_id')
        if room_id:
            room = ChatRoom.query.get_or_404(room_id)
            db.session.delete(room)
            db.session.commit()
            flash("Чат-комната успешно удалена.", "success")
        else:
            flash("Не выбрано ни одной комнаты для удаления.", "error")
    return redirect(url_for('admin_chat_rooms'))


# CANBAN TASKS


@app.route('/tasks')
def tasks():
    tasks = Task.query.all()  # Получаем все задачи с их статусами
    return render_template('tasks.html', tasks=tasks, title='Задачи')


# Создание задачи
@app.route('/tasks/create', methods=['GET', 'POST'])
def create_tasks():
    if request.method == 'POST':
        title = request.form['title']
        description = request.form['description']
        deadline = request.form['deadline']
        status_id = request.form['status_id']
        new_task = Task(title=title, description=description, deadline=deadline, status_id=status_id)
        db.session.add(new_task)
        db.session.commit()
        flash('Задача успешно создана!', 'success')
        return redirect(url_for('tasks'))
    statuses = Status.query.all()
    return render_template('create_tasks.html', statuses=statuses)


@app.route('/add-task', methods=['POST'])
def add_task():
    data = request.get_json()
    if not data or not data.get('title') or not data.get('status'):
        return jsonify({'success': False, 'message': 'Некорректные данные'}), 400

    title = data['title']
    status_name = data['status']

    # Преобразование статуса из имени в объект
    status = Status.query.filter_by(name=status_name).first()
    if not status:
        return jsonify({'success': False, 'message': 'Статус не найден'}), 400

    # Создание новой задачи
    task = Task(title=title, status_id=status.id)
    db.session.add(task)
    db.session.commit()

    return jsonify({'success': True, 'task': {'id': task.id, 'title': task.title, 'status': status.name}})


# Обновление задачи
@app.route('/tasks/update/<int:id>', methods=['GET', 'POST'])
def update_tasks(id):
    task = Task.query.get(id)
    if request.method == 'POST':
        task.status_id = request.form['status_id']
        db.session.commit()
        flash('Задача обновлена!', 'success')
        return redirect(url_for('tasks'))
    statuses = Status.query.all()
    return render_template('update_tasks.html', task=task, statuses=statuses)


# Удаление задачи
@app.route('/tasks/delete/<int:id>', methods=['GET', 'POST'])
def delete_tasks(id):
    task = Task.query.get(id)
    db.session.delete(task)
    db.session.commit()
    flash('Задача удалена!', 'success')
    return redirect(url_for('tasks'))


@app.route('/update_task_status/<int:task_id>', methods=['POST'])
@login_required
def update_task_status(task_id):
    status_data = request.get_json()
    new_status = status_data.get('status')  # Получаем статус из запроса

    # Карта сопоставления английских и русских названий статусов
    status_mapping = {
        'not-started': 'Не начато',
        'in-progress': 'В процессе',
        'completed': 'Выполнено',
    }

    # Проверяем, что новый статус валиден
    if new_status not in status_mapping:
        return jsonify({'error': 'Invalid status'}), 400

    # Преобразуем статус в русский
    localized_status = status_mapping[new_status]

    # Находим задачу и обновляем статус
    task = Task.query.get_or_404(task_id)
    status = Status.query.filter_by(name=localized_status).first()
    if not status:
        return jsonify({'error': 'Status not found'}), 404

    task.status_id = status.id
    db.session.commit()

    return jsonify({'message': 'Task status updated successfully'}), 200


# EXCEL

from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from io import BytesIO


@app.route('/export_projects', methods=['GET'])
@login_required
def export_projects():
    # Создание Excel-файла
    wb = Workbook()
    ws = wb.active
    ws.title = "Проекты"

    # Данные для экспорта (запрос к БД)
    projects = Project.query.all()

    # Стили для заголовков
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="447aaf")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                    bottom=Side(style='thin'))

    # Заполняем заголовки таблицы
    headers = ["ID", "Название", "Описание", "Дата начала", "Дата окончания", "Статус", "Приоритет", "Управляющий",
               "Ответственный"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Заполняем строки таблицы
    for row_num, project in enumerate(projects, start=2):
        ws.cell(row=row_num, column=1, value=project.id).border = border
        ws.cell(row=row_num, column=2, value=project.title).border = border
        ws.cell(row=row_num, column=3, value=project.description).border = border
        ws.cell(row=row_num, column=4,
                value=project.start_date.strftime('%d.%m.%Y') if project.start_date else "").border = border
        ws.cell(row=row_num, column=5,
                value=project.end_date.strftime('%d.%m.%Y') if project.end_date else "").border = border
        ws.cell(row=row_num, column=6, value=project.status.name if project.status else "Без статуса").border = border
        ws.cell(row=row_num, column=7, value=project.priority.level if project.priority else "Не задан").border = border
        ws.cell(row=row_num, column=8,
                value=f"{project.manager.first_name} {project.manager.last_name}").border = border
        ws.cell(row=row_num, column=9,
                value=f"{project.responsible_user.first_name} {project.responsible_user.last_name}" if project.responsible_user else "Не назначен").border = border

    # Автоширина столбцов
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Получаем букву столбца
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width

    # Сохранение в буфер
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Отправка файла пользователю
    return send_file(output, as_attachment=True, download_name="projects.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
